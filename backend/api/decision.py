from fastapi import APIRouter, HTTPException
from pydantic import BaseModel, Field
from typing import Dict, List, Optional

from services.decision_service import estimate_quotas

router = APIRouter()


class RoomSchema(BaseModel):
    min_supervisors: Optional[int] = Field(None, ge=0)


class SlotSchema(BaseModel):
    date: Optional[str]
    rooms: List[RoomSchema] = []


class DecisionRequest(BaseModel):
    counts_per_grade: Dict[str, int]
    planning: List[SlotSchema]
    default_min_per_room: Optional[int] = Field(1, ge=0)
    absence_majoration_pct: Optional[float] = Field(10.0, ge=0)
    min_diff_PR_MA: Optional[int] = Field(3, ge=0)
    max_quota_per_teacher: Optional[int] = Field(10, ge=1)


@router.post("/decision/estimate")
def estimate(request: DecisionRequest):
    try:
        res = estimate_quotas(
            counts_per_grade=request.counts_per_grade,
            planning=[s.dict() for s in request.planning],
            default_min_per_room=request.default_min_per_room,
            absence_majoration_pct=request.absence_majoration_pct,
            min_diff_PR_MA=request.min_diff_PR_MA,
            max_quota_per_teacher=request.max_quota_per_teacher,
        )
        return res
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from services.decision_service import DecisionService
from pydantic import BaseModel, Field
from typing import Dict, List, Optional
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

router = APIRouter(prefix="/decision", tags=["Aide à la Décision"])


class DecisionRequest(BaseModel):
    """Paramètres pour le calcul des recommandations"""

    min_surveillants_par_salle: int = Field(
        default=2, ge=1, le=5, description="Nombre minimum de surveillants par salle"
    )
    majoration_absences: float = Field(
        default=1.1,
        ge=1.0,
        le=1.5,
        description="Coefficient de majoration pour absences (1.1 = +10%)",
    )
    quota_min_groupe1: int = Field(
        default=4,
        ge=1,
        le=10,
        description="Quota minimal pour le groupe 1 (PR/MC/V)",
    )
    difference_min_pr_ma: int = Field(
        default=1,
        ge=1,
        le=5,
        description="Différence minimale entre PR/MC/V et MA",
    )
    difference_min_ma_as: int = Field(
        default=1,
        ge=1,
        le=5,
        description="Différence minimale entre MA et AS",
    )
    difference_min_as_ac: int = Field(
        default=1,
        ge=1,
        le=5,
        description="Différence minimale entre AS et AC/PES/PTC",
    )
    expert_quota: int = Field(
        default=3, ge=1, le=10, description="Quota fixe pour les experts"
    )


class DecisionResponse(BaseModel):
    """Réponse contenant les recommandations"""

    statistiques_globales: Dict
    quotas_recommandes: Dict
    voeux_autorises: Dict
    faisabilite: Dict
    distribution_temporelle: Dict  # Analyse de la distribution dans le temps
    alertes: List[Dict]
    parametres: Dict


@router.post("/calculer-recommandations", response_model=DecisionResponse)
def calculer_recommandations(
    request: DecisionRequest, db: Session = Depends(get_db)
):
    """
    📊 CALCULE LES RECOMMANDATIONS POUR LA CONFIGURATION DU PLANNING

    Ce endpoint analyse la situation actuelle (enseignants, examens) et produit:

    1. **Quotas recommandés par grade** selon la hiérarchie:
       - PR, MC, V: même quota (le plus bas)
       - MA: quota supérieur à PR/MC/V
       - AS: quota supérieur à MA
       - AC, PES, PTC: quota supérieur à AS
       - EX: quota fixe configurable

    2. **Nombre de créneaux de non-souhaits autorisés** par grade:
       - Plus le quota est élevé, moins l'enseignant peut exprimer de non-souhaits
       - Calcul automatique pour garantir la faisabilité

    3. **Analyse de faisabilité**:
       - OPTIMAL: Large marge pour gérer les absences
       - ACCEPTABLE: Marge faible mais suffisante
       - CRITIQUE: Ressources insuffisantes

    4. **Alertes et recommandations**:
       - Alertes critiques si problèmes détectés
       - Recommandations d'actions à entreprendre

    **Paramètres configurables:**
    - `min_surveillants_par_salle`: Nombre minimum de surveillants par salle (1-5)
    - `majoration_absences`: Coefficient pour absences (1.0-1.5, ex: 1.1 = +10%)
    - `quota_min_groupe1`: Quota minimal pour PR/MC/V (1-10, défaut: 3)
    - `difference_min_pr_ma`: Écart minimal entre PR/MC/V et MA (1-5)
    - `difference_min_ma_as`: Écart minimal entre MA et AS (1-5)
    - `difference_min_as_ac`: Écart minimal entre AS et AC/PES/PTC (1-5)
    - `expert_quota`: Quota fixe pour les experts (1-10)

    **Exemple d'utilisation:**
    ```json
    {
      "min_surveillants_par_salle": 2,
      "majoration_absences": 1.1,
      "quota_min_groupe1": 3,
      "difference_min_pr_ma": 1,
      "difference_min_ma_as": 1,
      "difference_min_as_ac": 1,
      "expert_quota": 3
    }
    ```

    **Retour:**
    - Statistiques globales (enseignants, examens, séances)
    - Quotas recommandés par grade
    - Créneaux de non-souhaits autorisés par grade
    - Analyse de faisabilité (statut, marges)
    - Liste d'alertes et recommandations
    """
    try:
        decision_service = DecisionService(db)

        recommandations = decision_service.calculer_recommandations(
            min_surveillants_par_salle=request.min_surveillants_par_salle,
            majoration_absences=request.majoration_absences,
            quota_min_groupe1=request.quota_min_groupe1,
            difference_min_pr_ma=request.difference_min_pr_ma,
            difference_min_ma_as=request.difference_min_ma_as,
            difference_min_as_ac=request.difference_min_as_ac,
            expert_quota=request.expert_quota,
        )

        return DecisionResponse(**recommandations)

    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Erreur lors du calcul des recommandations: {str(e)}"
        )


@router.post("/appliquer-quotas")
def appliquer_quotas(quotas: Dict[str, int], db: Session = Depends(get_db)):
    """
    💾 APPLIQUE LES QUOTAS RECOMMANDÉS DANS LA BASE DE DONNÉES

    Met à jour la configuration des quotas par grade dans la table `grade_config`.

    **Paramètres:**
    - `quotas`: Dictionnaire {grade_code: quota}

    **Exemple:**
    ```json
    {
      "PR": 4,
      "MC": 4,
      "MA": 5,
      "AS": 6,
      "AC": 7,
      "PTC": 7,
      "PES": 7,
      "EX": 3,
      "V": 4
    }
    ```

    **Retour:**
    - Nombre de grades mis à jour
    - Détails des modifications
    """
    from models.models import GradeConfig
    from config import GRADES

    try:
        modifications = []
        nb_updates = 0

        for grade_code, quota in quotas.items():
            # Vérifier que le grade existe dans la config
            if grade_code not in GRADES:
                continue

            # Chercher ou créer la configuration du grade
            grade_config = (
                db.query(GradeConfig)
                .filter(GradeConfig.grade_code == grade_code)
                .first()
            )

            ancien_quota = grade_config.nb_surveillances if grade_config else None

            if grade_config:
                # Mise à jour
                grade_config.nb_surveillances = quota
            else:
                # Création
                grade_config = GradeConfig(
                    grade_code=grade_code,
                    grade_nom=GRADES[grade_code]["nom"],
                    nb_surveillances=quota,
                )
                db.add(grade_config)

            modifications.append(
                {
                    "grade_code": grade_code,
                    "grade_nom": GRADES[grade_code]["nom"],
                    "ancien_quota": ancien_quota,
                    "nouveau_quota": quota,
                    "action": "mise_a_jour" if ancien_quota else "creation",
                }
            )

            nb_updates += 1

        db.commit()

        return {
            "success": True,
            "message": f"{nb_updates} grade(s) mis à jour avec succès",
            "nb_updates": nb_updates,
            "modifications": modifications,
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500, detail=f"Erreur lors de l'application des quotas: {str(e)}"
        )


@router.get("/quotas-actuels")
def obtenir_quotas_actuels(db: Session = Depends(get_db)):
    """
    📋 RÉCUPÈRE LES QUOTAS ACTUELLEMENT CONFIGURÉS

    Retourne la configuration actuelle des quotas par grade depuis la base de données.

    **Retour:**
    - Quotas par grade
    - Date de dernière modification
    """
    from models.models import GradeConfig
    from config import GRADES

    try:
        configs = db.query(GradeConfig).all()

        quotas = {}
        for config in configs:
            quotas[config.grade_code] = {
                "grade_nom": config.grade_nom,
                "quota": config.nb_surveillances,
                "created_at": config.created_at.isoformat() if config.created_at else None,
            }

        # Ajouter les grades qui n'ont pas encore de config
        for grade_code, grade_info in GRADES.items():
            if grade_code not in quotas:
                quotas[grade_code] = {
                    "grade_nom": grade_info["nom"],
                    "quota": grade_info["nb_surveillances"],
                    "created_at": None,
                    "source": "config_par_defaut",
                }

        return {"quotas": quotas, "nb_grades": len(quotas)}

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la récupération des quotas: {str(e)}",
        )


@router.post("/quotas-individuels")
def modifier_quota_individuel(
    enseignant_id: int, nouveau_quota: int, db: Session = Depends(get_db)
):
    """
    👤 MODIFIE LE QUOTA D'UN ENSEIGNANT SPÉCIFIQUE

    Permet d'ajuster manuellement le quota d'un enseignant particulier.

    **Paramètres:**
    - `enseignant_id`: ID de l'enseignant
    - `nouveau_quota`: Nouveau quota de surveillances

    **Note:** Cette modification est individuelle et écrase le quota par grade
    pour cet enseignant spécifique.
    """
    from models.models import Enseignant

    try:
        enseignant = (
            db.query(Enseignant).filter(Enseignant.id == enseignant_id).first()
        )

        if not enseignant:
            raise HTTPException(status_code=404, detail="Enseignant introuvable")

        # Note: Pour implémenter cette fonctionnalité, il faudrait ajouter
        # un champ `quota_individuel` dans la table Enseignant
        # Pour l'instant, on retourne un message informatif

        return {
            "success": False,
            "message": "Fonctionnalité à implémenter: ajout d'un champ quota_individuel dans la table Enseignant",
            "enseignant": {
                "id": enseignant.id,
                "nom": enseignant.nom,
                "prenom": enseignant.prenom,
                "grade_code": enseignant.grade_code,
            },
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la modification du quota individuel: {str(e)}",
        )


@router.post("/exporter-voeux-autorises")
def exporter_voeux_autorises(request: DecisionRequest, db: Session = Depends(get_db)):
    """
    📊 EXPORTE LES CRÉNEAUX DE NON-SOUHAITS AUTORISÉS AU FORMAT EXCEL

    Génère un fichier Excel contenant les informations sur les créneaux de non-souhaits
    autorisés par grade, basé sur les recommandations calculées.

    **Colonnes du fichier:**
    - Code Grade
    - Nom du Grade
    - Créneaux Autorisés

    **Paramètres:** Les mêmes que pour `/calculer-recommandations`

    **Retour:** Fichier Excel (.xlsx) nommé `creneaux_non_souhaits_autorises.xlsx`
    """
    try:
        decision_service = DecisionService(db)

        # Calculer les recommandations
        recommandations = decision_service.calculer_recommandations(
            min_surveillants_par_salle=request.min_surveillants_par_salle,
            majoration_absences=request.majoration_absences,
            quota_min_groupe1=request.quota_min_groupe1,
            difference_min_pr_ma=request.difference_min_pr_ma,
            difference_min_ma_as=request.difference_min_ma_as,
            difference_min_as_ac=request.difference_min_as_ac,
            expert_quota=request.expert_quota,
        )

        # Créer un fichier Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Créneaux Non-Souhaits"

        # Style pour l'en-tête
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # En-têtes
        headers = ["Code Grade", "Nom du Grade", "Créneaux Autorisés"]
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Données
        voeux_autorises = recommandations.get("voeux_autorises", {})
        quotas_recommandes = recommandations.get("quotas_recommandes", {})
        
        row_num = 2
        for grade_code, info in voeux_autorises.items():
            # Calculer les voeux autorisés ajustés
            quota_actuel = quotas_recommandes.get(grade_code, {}).get("quota", 0)
            nb_total_seances = info.get("nb_total_seances", 0)
            
            # Formule stricte : max(0, floor((nb_total_seances - quota_actuel) * 0.6))
            difference = nb_total_seances - quota_actuel
            nb_voeux_max_recommande = max(0, int(difference * 0.6))
            
            sheet.cell(row=row_num, column=1).value = grade_code
            sheet.cell(row=row_num, column=2).value = info.get("grade_nom", "")
            sheet.cell(row=row_num, column=3).value = nb_voeux_max_recommande
            
            # Aligner les données
            for col in range(1, 4):
                cell = sheet.cell(row=row_num, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1

        # Ajuster la largeur des colonnes
        column_widths = [15, 30, 20]
        for col_num, width in enumerate(column_widths, 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width

        # Sauvegarder dans un buffer
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        # Retourner le fichier
        headers = {
            'Content-Disposition': 'attachment; filename="creneaux_non_souhaits_autorises.xlsx"'
        }
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de l'exportation: {str(e)}"
        )


@router.post("/importer-exceptions")
async def importer_exceptions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """
    📥 IMPORTE LES EXCEPTIONS D'ENSEIGNANTS DEPUIS UN FICHIER EXCEL

    Permet d'importer un fichier Excel contenant les exceptions d'enseignants
    basées sur les absences. Le quota de chaque enseignant sera ajusté selon
    les absences indiquées.

    ⚠️ **ATTENTION** : Cette opération supprime toutes les exceptions existantes
    avant d'importer les nouvelles. Seuls les enseignants listés dans le fichier
    auront des exceptions après l'import.

    **Colonnes attendues:**
    - Nom: Nom de famille de l'enseignant
    - Prénom: Prénom de l'enseignant
    - Code: Code de l'enseignant (code_smartex)
    - Absences: Nombre d'absences (positif ou négatif)

    **Règles d'ajustement:**
    - Si Absences > 0: L'enseignant est marqué comme exception et son quota
      est augmenté de ce nombre (quota_grade + absences)
    - Si Absences < 0: L'enseignant est marqué comme exception et son quota
      est diminué de ce nombre (quota_grade - |absences|)
    - Si Absences = 0 ou vide: L'enseignant n'est pas marqué comme exception

    **Exemple:**
    ```
    Nom       | Prénom | Code | Absences
    Belhouene | Imen   | 100  | 1
    Dupont    | Jean   | 101  | -2
    ```

    **Retour:**
    - Nombre d'exceptions importées
    - Liste des erreurs éventuelles
    """
    from config import UPLOAD_DIR
    import os
    import shutil

    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=400,
            detail="Le fichier doit être au format Excel (.xlsx ou .xls)"
        )

    file_path = os.path.join(UPLOAD_DIR, file.filename)

    try:
        # Sauvegarder temporairement le fichier
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        # Traiter le fichier
        decision_service = DecisionService(db)
        count, erreurs = decision_service.importer_exceptions_absences(file_path)

        # Nettoyer
        os.remove(file_path)

        return {
            "success": True,
            "message": f"{count} exceptions importées avec succès",
            "nb_importes": count,
            "erreurs": erreurs
        }

    except Exception as e:
        if os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de l'import: {str(e)}"
        )


@router.delete("/supprimer-exceptions")
def supprimer_exceptions(db: Session = Depends(get_db)):
    """
    🗑️  SUPPRIME TOUTES LES EXCEPTIONS D'ENSEIGNANTS

    Remet tous les enseignants marqués comme exceptions à l'état normal.
    Les quotas exceptionnels sont supprimés et les quotas de grade sont rétablis.

    **Retour:**
    - Nombre d'exceptions supprimées
    """
    try:
        decision_service = DecisionService(db)
        nb_supprimes = decision_service.supprimer_exceptions()
        
        return {
            "success": True,
            "message": f"{nb_supprimes} exception(s) supprimée(s) avec succès",
            "nb_supprimes": nb_supprimes
        }
    
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la suppression: {str(e)}"
        )


@router.get("/enseignants-exceptions")
def obtenir_enseignants_exceptions(db: Session = Depends(get_db)):
    """
    📋 RÉCUPÈRE LA LISTE DES ENSEIGNANTS AVEC EXCEPTIONS

    Retourne tous les enseignants marqués comme exceptions avec leurs quotas personnalisés.

    **Retour:**
    - Liste des enseignants avec exceptions
    - Détails : nom, prénom, code, grade, quota_grade, quota_exception
    """
    from models.models import Enseignant, GradeConfig
    from config import GRADES

    try:
        enseignants_exceptions = db.query(Enseignant).filter(
            Enseignant.is_Exception == True
        ).all()

        resultats = []
        for ens in enseignants_exceptions:
            # Récupérer le quota du grade
            grade_config = db.query(GradeConfig).filter(
                GradeConfig.grade_code == ens.grade_code
            ).first()
            
            if grade_config:
                quota_grade = grade_config.nb_surveillances
            elif ens.grade_code in GRADES:
                quota_grade = GRADES[ens.grade_code]["nb_surveillances"]
            else:
                quota_grade = 0
            
            difference = ens.quota_Exception - quota_grade if ens.quota_Exception is not None else 0
            
            resultats.append({
                "id": ens.id,
                "nom": ens.nom,
                "prenom": ens.prenom,
                "code_smartex": ens.code_smartex,
                "grade_code": ens.grade_code,
                "grade_nom": ens.grade,
                "quota_grade": quota_grade,
                "quota_exception": ens.quota_Exception,
                "difference": difference,
                "type_exception": "augmentation" if difference > 0 else "diminution" if difference < 0 else "egal"
            })

        return {
            "success": True,
            "nb_exceptions": len(resultats),
            "exceptions": resultats
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Erreur lors de la récupération des exceptions: {str(e)}"
        )
